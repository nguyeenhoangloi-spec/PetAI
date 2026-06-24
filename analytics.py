# analytics.py
# Blueprint cho thống kê nhận diện
from flask import Blueprint, render_template, session, redirect, url_for, flash, jsonify, request
from connect import get_connection
from models import PredictionHistory

stats_bp = Blueprint("statistics", __name__)


@stats_bp.route("/")
def statistics():
    """Trang thống kê"""
    user_id_any = session.get("user_id")
    if user_id_any is None:
        flash("Vui lòng đăng nhập để xem thống kê.", "warning")
        return redirect(url_for("login.login"))
    conn = None
    try:
        from datetime import datetime, timedelta

        selected_days = request.args.get("days", "30", type=str)
        if selected_days not in ("7", "30", "90", "0", "custom"):
            selected_days = "30"

        start_date_str = request.args.get("start_date", "")
        end_date_str = request.args.get("end_date", "")
        
        start_at = None
        end_at = None
        start_prev = None
        end_prev = None
        has_comparison = False
        duration_days = 30

        now = datetime.now()

        if selected_days == "7":
            start_at = now - timedelta(days=7)
            end_at = now
            start_prev = now - timedelta(days=14)
            end_prev = now - timedelta(days=7)
            has_comparison = True
            duration_days = 7
        elif selected_days == "30":
            start_at = now - timedelta(days=30)
            end_at = now
            start_prev = now - timedelta(days=60)
            end_prev = now - timedelta(days=30)
            has_comparison = True
            duration_days = 30
        elif selected_days == "90":
            start_at = now - timedelta(days=90)
            end_at = now
            start_prev = now - timedelta(days=180)
            end_prev = now - timedelta(days=90)
            has_comparison = True
            duration_days = 90
        elif selected_days == "custom":
            try:
                start_at = datetime.strptime(start_date_str, "%Y-%m-%d")
                end_at = datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                duration_days = (end_at - start_at).days + 1
                if duration_days < 1:
                    duration_days = 1
                start_prev = start_at - timedelta(days=duration_days)
                end_prev = start_at - timedelta(seconds=1)
                has_comparison = True
            except Exception:
                selected_days = "30"
                start_at = now - timedelta(days=30)
                end_at = now
                start_prev = now - timedelta(days=60)
                end_prev = now - timedelta(days=30)
                has_comparison = True
                duration_days = 30
        else:
            start_at = None
            end_at = None
            has_comparison = False

        conn = get_connection()
        user_id = int(user_id_any)

        # Get current stats
        stats = PredictionHistory.get_stats(conn, user_id, start_at=start_at, end_at=end_at)
        
        # Get latest image path for each top breed to display as thumbnail
        if stats.get('top_breeds') and conn:
            with conn.cursor() as cur:
                for item in stats['top_breeds']:
                    cur.execute("""
                        SELECT image_path 
                        FROM prediction_history 
                        WHERE user_id = %s AND breed = %s AND image_path IS NOT NULL AND image_path != ''
                        ORDER BY created_at DESC LIMIT 1
                    """, (user_id, item['breed']))
                    row = cur.fetchone()
                    item['image_path'] = row[0] if row else None
        
        # Get unique breed count
        unique_breed_count = PredictionHistory.get_unique_breed_count(conn, user_id, start_at=start_at, end_at=end_at)
        
        # Get daily counts
        if selected_days == "0":
            daily_counts = PredictionHistory.get_daily_counts(conn, user_id, days=None)
        else:
            daily_counts = PredictionHistory.get_daily_counts(conn, user_id, start_at=start_at, end_at=end_at)
            
        # Get confidence distribution
        confidence_dist = PredictionHistory.get_confidence_distribution(conn, user_id, start_at=start_at, end_at=end_at)

        # Recent predictions (limit 10)
        recent_predictions = PredictionHistory.get_by_user(conn, user_id, limit=10)
        
        # Calculate comparison badges
        scans_diff_pct = None
        conf_diff = None
        new_breeds_count = 0
        
        if has_comparison and start_prev and end_prev:
            prev_stats = PredictionHistory.get_stats(conn, user_id, start_at=start_prev, end_at=end_prev)
            
            # Scans diff %
            curr_scans = stats.get('total_predictions', 0)
            prev_scans = prev_stats.get('total_predictions', 0)
            if prev_scans > 0:
                scans_diff_pct = round(((curr_scans - prev_scans) / prev_scans) * 100, 1)
            elif curr_scans > 0:
                scans_diff_pct = 100.0
            else:
                scans_diff_pct = 0.0
                
            # Avg confidence diff
            curr_conf = stats.get('avg_confidence', 0.0) * 100
            prev_conf = prev_stats.get('avg_confidence', 0.0) * 100
            conf_diff = round(curr_conf - prev_conf, 1)
            
            # New breeds count
            with conn.cursor() as cur:
                if user_id is not None:
                    cur.execute("""
                        SELECT DISTINCT breed 
                        FROM prediction_history 
                        WHERE user_id = %s AND created_at < %s AND breed IS NOT NULL AND breed != ''
                    """, (user_id, start_at))
                    breeds_before = {row[0] for row in cur.fetchall()}
                    
                    cur.execute("""
                        SELECT DISTINCT breed 
                        FROM prediction_history 
                        WHERE user_id = %s AND created_at >= %s AND created_at <= %s AND breed IS NOT NULL AND breed != ''
                    """, (user_id, start_at, end_at))
                    breeds_current = {row[0] for row in cur.fetchall()}
                else:
                    cur.execute("""
                        SELECT DISTINCT breed 
                        FROM prediction_history 
                        WHERE created_at < %s AND breed IS NOT NULL AND breed != ''
                    """, (start_at,))
                    breeds_before = {row[0] for row in cur.fetchall()}
                    
                    cur.execute("""
                        SELECT DISTINCT breed 
                        FROM prediction_history 
                        WHERE created_at >= %s AND created_at <= %s AND breed IS NOT NULL AND breed != ''
                    """, (start_at, end_at))
                    breeds_current = {row[0] for row in cur.fetchall()}
            
            new_breeds = breeds_current - breeds_before
            new_breeds_count = len(new_breeds)

        # Get latest scan relative time
        latest_prediction = PredictionHistory.get_by_user(conn, user_id, limit=1)
        latest_scan_time_vi = "Chưa có hoạt động"
        latest_scan_time_en = "No activity"
        if latest_prediction:
            last_dt = latest_prediction[0]['created_at']
            if last_dt:
                time_str = last_dt.strftime("%H:%M")
                date_str = last_dt.strftime("%d/%m/%Y")
                latest_scan_time_vi = f"Hôm nay {time_str}" if last_dt.date() == now.date() else (f"Hôm qua {time_str}" if last_dt.date() == (now - timedelta(days=1)).date() else f"{date_str} {time_str}")
                latest_scan_time_en = f"Today {time_str}" if last_dt.date() == now.date() else (f"Yesterday {time_str}" if last_dt.date() == (now - timedelta(days=1)).date() else f"{date_str} {time_str}")

        # Compute insights / highlights
        most_common_breed = "Chưa có dữ liệu"
        if stats.get('top_breeds'):
            most_common_breed = stats['top_breeds'][0]['breed']
            
        confidence_brackets = ["0-20%", "20-40%", "40-60%", "60-80%", "80-100%"]
        max_bracket_index = -1
        max_bracket_count = -1
        for i, val in enumerate(confidence_dist):
            if val > max_bracket_count:
                max_bracket_count = val
                max_bracket_index = i
        peak_confidence_bracket = confidence_brackets[max_bracket_index] if max_bracket_count > 0 else "Chưa có dữ liệu"

        peak_active_date = "Chưa có dữ liệu"
        peak_active_weekday_vi = ""
        peak_active_weekday_en = ""
        max_daily_count = -1
        for day in daily_counts:
            if day['count'] > max_daily_count:
                max_daily_count = day['count']
                peak_active_date = day['date']
        if max_daily_count <= 0:
            peak_active_date = "Chưa có dữ liệu"
        else:
            try:
                # Parse weekday from format dd/mm
                if peak_active_date and '/' in peak_active_date:
                    parts = peak_active_date.split('/')
                    if len(parts) == 2:
                        day_val = int(parts[0])
                        month_val = int(parts[1])
                        # Check if month is valid (not monthly format like mm/yy)
                        if month_val <= 12:
                            year_val = datetime.now().year
                            dt_obj = datetime(year_val, month_val, day_val)
                            weekdays_vi = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                            weekdays_en = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                            peak_active_weekday_vi = weekdays_vi[dt_obj.weekday()]
                            peak_active_weekday_en = weekdays_en[dt_obj.weekday()]
            except Exception:
                pass

        return render_template(
            "statistics.html",
            stats=stats,
            recent_predictions=recent_predictions,
            unique_breed_count=unique_breed_count,
            daily_counts=daily_counts,
            confidence_dist=confidence_dist,
            selected_days=selected_days,
            start_date=start_date_str or (start_at.strftime("%Y-%m-%d") if start_at else ""),
            end_date=end_date_str or (end_at.strftime("%Y-%m-%d") if end_at else ""),
            scans_diff_pct=scans_diff_pct,
            conf_diff=conf_diff,
            new_breeds_count=new_breeds_count,
            latest_scan_time_vi=latest_scan_time_vi,
            latest_scan_time_en=latest_scan_time_en,
            most_common_breed=most_common_breed,
            peak_confidence_bracket=peak_confidence_bracket,
            peak_active_date=peak_active_date,
            peak_active_weekday_vi=peak_active_weekday_vi,
            peak_active_weekday_en=peak_active_weekday_en,
            duration_days=duration_days
        )
    except Exception as e:
        print(f"Error loading statistics: {e}")
        flash("Không thể tải thống kê. Vui lòng thử lại.", "error")
        return redirect(url_for("dashboard.dashboard"))
    finally:
        if conn:
            conn.close()


@stats_bp.route("/api/stats")
def api_stats():
    """API lấy thống kê"""
    user_id_any = session.get("user_id")
    if user_id_any is None:
        return jsonify({"error": "Not authenticated"}), 401
    conn = None
    try:
        conn = get_connection()
        user_id = int(user_id_any)
        stats = PredictionHistory.get_stats(conn, user_id)
        return jsonify(stats)
    except Exception as e:
        print(f"Error in API: {e}")
        return jsonify({"error": "Internal error"}), 500
    finally:
        if conn:
            conn.close()


@stats_bp.route("/export")
def export_statistics():
    """Xuất báo cáo thống kê dưới dạng xlsx hoặc csv"""
    user_id_any = session.get("user_id")
    if user_id_any is None:
        return jsonify({"error": "Not authenticated"}), 401

    from datetime import datetime, timedelta
    import io

    fmt = request.args.get("format", "xlsx").lower()
    if fmt not in ("xlsx", "csv"):
        fmt = "xlsx"

    selected_days = request.args.get("days", "30", type=str)
    start_date_str = request.args.get("start_date", "")
    end_date_str = request.args.get("end_date", "")

    now = datetime.now()
    start_at = None
    end_at = None

    if selected_days == "7":
        start_at = now - timedelta(days=7)
        end_at = now
    elif selected_days == "30":
        start_at = now - timedelta(days=30)
        end_at = now
    elif selected_days == "90":
        start_at = now - timedelta(days=90)
        end_at = now
    elif selected_days == "custom":
        try:
            start_at = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_at = datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except Exception:
            start_at = now - timedelta(days=30)
            end_at = now
    # selected_days == "0" => all time (start_at=None)

    conn = None
    try:
        conn = get_connection()
        user_id = int(user_id_any)

        # Determine language (i18n)
        default_lang = "vi"
        try:
            from models import SystemConfig, UserSettings
            default_lang = SystemConfig.get(conn, "default_lang", "vi")
        except Exception:
            pass

        lang = request.cookies.get("siteLanguage")
        if not lang or lang not in {"vi", "en"}:
            lang = default_lang
        if user_id_any is not None and (not request.cookies.get("siteLanguage")):
            try:
                user_settings = UserSettings.get_or_create(conn, int(user_id_any))
                lang_raw = (user_settings or {}).get("language")
                if lang_raw in {"vi", "en"}:
                    lang = lang_raw
            except Exception:
                pass
        if lang not in {"vi", "en"}:
            lang = "vi"

        # Translation dictionary
        T = {
            "vi": {
                "title": "📊 BÁO CÁO THỐNG KÊ PETAI",
                "exported_at": "Xuất ngày: {}",
                "summary_header": "== TÓM TẮT THỐNG KÊ ==",
                "metric": "Chỉ số",
                "value": "Giá trị",
                "total_scans": "Tổng lượt nhận diện",
                "avg_confidence": "Độ tin cậy trung bình (%)",
                "unique_breeds": "Số giống chó đã khám phá",
                
                "insights_header": "== THÔNG TIN NỔI BẬT ==",
                "most_common_breed": "Giống phổ biến nhất",
                "peak_confidence": "Khoảng tin cậy phổ biến",
                "comparison": "So với kỳ trước",
                "peak_active_day": "Ngày hoạt động nhiều nhất",
                
                "trends_header": "== XU HƯỚNG NHẬN DIỆN ==",
                "date": "Ngày",
                "scans_count": "Số lượt nhận diện",
                
                "top_5_header": "== TOP 5 GIỐNG PHỔ BIẾN NHẤT ==",
                "rank": "Hạng",
                "breed": "Giống chó",
                "pct": "Tỉ lệ (%)",
                
                "dist_header": "== PHÂN BỐ GIỐNG CHÓ ==",
                
                "conf_header": "== PHÂN BỐ ĐỘ TIN CẬY ==",
                "conf_interval": "Khoảng độ tin cậy",

                "recent_header": "== KẾT QUẢ NHẬN DIỆN GẦN ĐÂY ==",
                "time": "Thời gian",
                "species": "Loài",
                "conf": "Độ tin cậy",
                "species_dog": "Chó",
                
                "no_data": "Chưa có dữ liệu",
                "others": "Khác",
                "growth_label": "Tăng trưởng lượt quét",
                "sheet_summary": "Tóm tắt & Nổi bật",
                "sheet_trends": "Xu hướng nhận diện",
                "sheet_top5": "Top 5 giống phổ biến",
                "sheet_dist": "Phân bố giống chó",
                "sheet_conf": "Phân bố độ tin cậy",
                "sheet_recent": "Kết quả gần đây",
                "all_time": "Tất cả",
                "compared_to_days": "so với {} ngày trước",
                "times": "lần"
            },
            "en": {
                "title": "📊 PETAI STATISTICS REPORT",
                "exported_at": "Exported at: {}",
                "summary_header": "== STATISTICAL SUMMARY ==",
                "metric": "Metric",
                "value": "Value",
                "total_scans": "Total Scans",
                "avg_confidence": "Avg Confidence (%)",
                "unique_breeds": "Unique Breeds Explored",
                
                "insights_header": "== HIGHLIGHTED INSIGHTS ==",
                "most_common_breed": "Most Common Breed",
                "peak_confidence": "Peak Confidence Bracket",
                "comparison": "Compared to Last Period",
                "peak_active_day": "Peak Active Day",
                
                "trends_header": "== RECOGNITION TRENDS ==",
                "date": "Date",
                "scans_count": "Scans Count",
                
                "top_5_header": "== TOP 5 MOST POPULAR BREEDS ==",
                "rank": "Rank",
                "breed": "Dog Breed",
                "pct": "Percentage (%)",
                
                "dist_header": "== DOG BREED DISTRIBUTION ==",
                
                "conf_header": "== CONFIDENCE DISTRIBUTION ==",
                "conf_interval": "Confidence Interval",

                "recent_header": "== RECENT RECOGNITION RESULTS ==",
                "time": "Time",
                "species": "Species",
                "conf": "Confidence",
                "species_dog": "Dog",
                
                "no_data": "No data available",
                "others": "Others",
                "growth_label": "Scan growth",
                "sheet_summary": "Summary & Highlights",
                "sheet_trends": "Recognition Trends",
                "sheet_top5": "Top 5 Popular Breeds",
                "sheet_dist": "Breed Distribution",
                "sheet_conf": "Confidence Distribution",
                "sheet_recent": "Recent Results",
                "all_time": "All time",
                "compared_to_days": "compared to {} days ago",
                "times": "times"
            }
        }[lang]

        # Calculate dates for previous period comparison
        has_comparison = False
        duration_days = 30
        start_prev = None
        end_prev = None
        if selected_days == "7":
            start_prev = now - timedelta(days=14)
            end_prev = now - timedelta(days=7)
            has_comparison = True
            duration_days = 7
        elif selected_days == "30":
            start_prev = now - timedelta(days=60)
            end_prev = now - timedelta(days=30)
            has_comparison = True
            duration_days = 30
        elif selected_days == "90":
            start_prev = now - timedelta(days=180)
            end_prev = now - timedelta(days=90)
            has_comparison = True
            duration_days = 90
        elif selected_days == "custom":
            duration_days = (end_at - start_at).days + 1
            if duration_days < 1:
                duration_days = 1
            start_prev = start_at - timedelta(days=duration_days)
            end_prev = start_at - timedelta(seconds=1)
            has_comparison = True

        # Fetch stats & records
        stats = PredictionHistory.get_stats(conn, user_id, start_at=start_at, end_at=end_at)
        unique_breed_count = PredictionHistory.get_unique_breed_count(conn, user_id, start_at=start_at, end_at=end_at)
        daily_counts = PredictionHistory.get_daily_counts(
            conn, user_id,
            days=None,
            start_at=start_at,
            end_at=end_at
        )
        confidence_dist = PredictionHistory.get_confidence_distribution(conn, user_id, start_at=start_at, end_at=end_at)
        recent_predictions = PredictionHistory.get_by_user(conn, user_id, limit=5)

        top_breeds = stats.get("top_breeds", [])
        total_preds = stats.get("total_predictions", 0)
        avg_conf = round((stats.get("avg_confidence", 0) or 0) * 100, 1)

        # 1. Insights: Most common breed
        most_common_breed = T["no_data"]
        if top_breeds:
            raw_breed = top_breeds[0].get("breed", "")
            if raw_breed:
                if lang == "en":
                    from i18n_server import translate_breed_vi_to_en
                    most_common_breed = translate_breed_vi_to_en(raw_breed)
                else:
                    most_common_breed = raw_breed
                # Handle "Nghi lai" translation logic
                if ":" in most_common_breed:
                    most_common_breed = most_common_breed.split(":")[-1].strip()

        # 2. Insights: Peak confidence bracket
        conf_labels = ["0–20%", "20–40%", "40–60%", "60–80%", "80–100%"]
        max_bracket_index = -1
        max_bracket_count = -1
        for idx, val in enumerate(confidence_dist):
            if val > max_bracket_count:
                max_bracket_count = val
                max_bracket_index = idx
        peak_confidence_bracket = conf_labels[max_bracket_index] if max_bracket_count > 0 else T["no_data"]

        # 3. Insights: Growth comparison
        scans_diff_pct = None
        if has_comparison and start_prev and end_prev:
            prev_stats = PredictionHistory.get_stats(conn, user_id, start_at=start_prev, end_at=end_prev)
            curr_scans = total_preds
            prev_scans = prev_stats.get('total_predictions', 0)
            if prev_scans > 0:
                scans_diff_pct = round(((curr_scans - prev_scans) / prev_scans) * 100, 1)
            elif curr_scans > 0:
                scans_diff_pct = 100.0
            else:
                scans_diff_pct = 0.0

        if selected_days == "0":
            comparison_text = T["all_time"]
        elif scans_diff_pct is not None:
            comparison_text = f"+{scans_diff_pct}%" if scans_diff_pct > 0 else f"{scans_diff_pct}%"
            comparison_text += f" ({T['compared_to_days'].format(duration_days)})"
        else:
            comparison_text = "--"

        # 4. Insights: Peak active date & weekday
        peak_active_date_str = T["no_data"]
        max_daily_count = -1
        for day in daily_counts:
            if day['count'] > max_daily_count:
                max_daily_count = day['count']
                peak_active_date_str = day['date']
        if max_daily_count <= 0:
            peak_active_date_str = T["no_data"]
        else:
            try:
                if peak_active_date_str and '/' in peak_active_date_str:
                    parts = peak_active_date_str.split('/')
                    if len(parts) == 2:
                        day_val = int(parts[0])
                        month_val = int(parts[1])
                        if month_val <= 12:
                            year_val = datetime.now().year
                            dt_obj = datetime(year_val, month_val, day_val)
                            if lang == "vi":
                                weekdays = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                                peak_active_date_str = f"{peak_active_date_str} ({weekdays[dt_obj.weekday()]})"
                            else:
                                weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
                                peak_active_date_str = f"{peak_active_date_str} ({weekdays[dt_obj.weekday()]})"
            except Exception:
                pass

        # 5. Breed Distribution (Top 5 + Others)
        breed_dist = []
        sum_top5 = 0
        for b in top_breeds[:5]:
            sum_top5 += b.get("count", 0)
            breed_name = b.get("breed", "")
            if lang == "en":
                from i18n_server import translate_breed_vi_to_en
                breed_name = translate_breed_vi_to_en(breed_name)
            breed_dist.append({
                "breed": breed_name,
                "count": b.get("count", 0),
                "pct": round(b.get("count", 0) / total_preds * 100, 1) if total_preds else 0.0
            })
        others_count = total_preds - sum_top5
        if others_count > 0:
            breed_dist.append({
                "breed": T["others"],
                "count": others_count,
                "pct": round(others_count / total_preds * 100, 1) if total_preds else 0.0
            })

        # Period label for filename
        if selected_days == "0":
            period_label = "tatca" if lang == "vi" else "all"
        elif selected_days == "custom":
            period_label = f"{start_date_str}_to_{end_date_str}"
        else:
            period_label = f"{selected_days}ngay" if lang == "vi" else f"{selected_days}days"

        filename_base = f"thongke_petai_{period_label}_{now.strftime('%Y%m%d')}" if lang == "vi" else f"statistics_petai_{period_label}_{now.strftime('%Y%m%d')}"

        # ── CSV ──────────────────────────────────────────────────────────────
        if fmt == "csv":
            import csv as csv_mod
            output = io.StringIO()
            writer = csv_mod.writer(output)

            # Summary Section
            writer.writerow([T["summary_header"]])
            writer.writerow([T["metric"], T["value"]])
            writer.writerow([T["total_scans"], total_preds])
            writer.writerow([T["avg_confidence"], avg_conf])
            writer.writerow([T["unique_breeds"], unique_breed_count])
            writer.writerow([])

            # Highlighted Insights Section
            writer.writerow([T["insights_header"]])
            writer.writerow([T["metric"], T["value"]])
            writer.writerow([T["most_common_breed"], most_common_breed])
            writer.writerow([T["peak_confidence"], peak_confidence_bracket])
            writer.writerow([T["comparison"], comparison_text])
            writer.writerow([T["peak_active_day"], peak_active_date_str])
            writer.writerow([])

            # Daily counts (Trends) Section
            writer.writerow([T["trends_header"]])
            writer.writerow([T["date"], T["scans_count"]])
            for d in daily_counts:
                writer.writerow([d.get("date", ""), d.get("count", 0)])
            writer.writerow([])

            # Top 5 breeds Section
            writer.writerow([T["top_5_header"]])
            writer.writerow([T["rank"], T["breed"], T["scans_count"], T["pct"]])
            for idx, b in enumerate(top_breeds[:5], 1):
                pct = round(b.get("count", 0) / total_preds * 100, 1) if total_preds else 0.0
                breed_name = b.get("breed", "")
                if lang == "en":
                    from i18n_server import translate_breed_vi_to_en
                    breed_name = translate_breed_vi_to_en(breed_name)
                writer.writerow([idx, breed_name, b.get("count", 0), pct])
            writer.writerow([])

            # Breed distribution (Top 5 + Others) Section
            writer.writerow([T["dist_header"]])
            writer.writerow([T["breed"], T["scans_count"], T["pct"]])
            for b in breed_dist:
                writer.writerow([b["breed"], b["count"], b["pct"]])
            writer.writerow([])

            # Confidence distribution Section
            writer.writerow([T["conf_header"]])
            writer.writerow([T["conf_interval"], T["scans_count"]])
            for label, val in zip(conf_labels, confidence_dist):
                writer.writerow([label, val])
            writer.writerow([])

            # Recent results Section
            writer.writerow([T["recent_header"]])
            writer.writerow([T["time"], T["breed"], T["species"], T["conf"]])
            for p in recent_predictions:
                p_date = p["created_at"].strftime('%d/%m/%Y %H:%M') if p.get("created_at") else ""
                p_breed = p["breed"]
                if lang == "en":
                    from i18n_server import translate_breed_vi_to_en
                    p_breed = p.get("breed_en") or translate_breed_vi_to_en(p_breed)
                p_species = T["species_dog"] if p.get("species") in ("Dog", "Chó") else p.get("species", "")
                p_conf_val = f"{round((p.get('confidence') or 0) * 100)}%"
                writer.writerow([p_date, p_breed, p_species, p_conf_val])

            output.seek(0)
            from flask import Response
            return Response(
                "\ufeff" + output.getvalue(),  # BOM for Excel UTF-8
                mimetype="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"}
            )

        # ── XLSX ─────────────────────────────────────────────────────────────
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from flask import send_file

        wb = openpyxl.Workbook()

        # Design style variables
        header_fill = PatternFill("solid", fgColor="004AC6")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Inter")
        subheader_fill = PatternFill("solid", fgColor="EEF2FF")
        subheader_font = Font(bold=True, color="1E3A8A", size=11, name="Inter")
        thin_side = Side(style="thin", color="D1D5DB")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        def style_header(cell, text):
            cell.value = text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        def style_cell(cell, value, bold=False, align="left", fill=None):
            cell.value = value
            cell.font = Font(bold=bold, size=10, name="Inter")
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill

        # ── Sheet 1: Summary & Highlights ──
        ws1 = wb.active
        ws1.title = T["sheet_summary"]

        # Report Title Row
        ws1.merge_cells("A1:B1")
        ws1["A1"].value = T["title"]
        ws1["A1"].font = Font(bold=True, size=14, color="004AC6", name="Inter")
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 30

        # Export Date Row
        ws1.merge_cells("A2:B2")
        ws1["A2"].value = T["exported_at"].format(now.strftime('%d/%m/%Y %H:%M'))
        ws1["A2"].font = Font(italic=True, size=9, color="6B7280", name="Inter")
        ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[2].height = 20

        # 1. General Summary Table
        ws1.merge_cells("A4:B4")
        ws1["A4"].value = T["summary_header"].replace("==", "").strip()
        ws1["A4"].font = subheader_font
        ws1["A4"].fill = subheader_fill
        ws1["A4"].alignment = Alignment(horizontal="left", vertical="center")
        ws1["A4"].border = border
        ws1.row_dimensions[4].height = 24

        style_header(ws1["A5"], T["metric"])
        style_header(ws1["B5"], T["value"])
        ws1.row_dimensions[5].height = 20

        summary_rows = [
            (T["total_scans"], total_preds),
            (T["avg_confidence"], avg_conf),
            (T["unique_breeds"], unique_breed_count),
        ]
        for idx, (k, v) in enumerate(summary_rows, 6):
            style_cell(ws1.cell(row=idx, column=1), k, bold=True)
            style_cell(ws1.cell(row=idx, column=2), v, align="center")
            ws1.row_dimensions[idx].height = 20

        # 2. Highlighted Insights Table
        start_row_ins = 11
        ws1.merge_cells(start_row=start_row_ins, start_column=1, end_row=start_row_ins, end_column=2)
        ws1.cell(row=start_row_ins, column=1).value = T["insights_header"].replace("==", "").strip()
        ws1.cell(row=start_row_ins, column=1).font = subheader_font
        ws1.cell(row=start_row_ins, column=1).fill = subheader_fill
        ws1.cell(row=start_row_ins, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws1.cell(row=start_row_ins, column=1).border = border
        ws1.row_dimensions[start_row_ins].height = 24

        style_header(ws1.cell(row=start_row_ins + 1, column=1), T["metric"])
        style_header(ws1.cell(row=start_row_ins + 1, column=2), T["value"])
        ws1.row_dimensions[start_row_ins + 1].height = 20

        insight_rows = [
            (T["most_common_breed"], most_common_breed),
            (T["peak_confidence"], peak_confidence_bracket),
            (T["comparison"], comparison_text),
            (T["peak_active_day"], peak_active_date_str),
        ]
        for idx, (k, v) in enumerate(insight_rows, start_row_ins + 2):
            style_cell(ws1.cell(row=idx, column=1), k, bold=True)
            style_cell(ws1.cell(row=idx, column=2), v, align="center")
            ws1.row_dimensions[idx].height = 20

        ws1.column_dimensions["A"].width = 32
        ws1.column_dimensions["B"].width = 32

        # ── Sheet 2: Recognition Trends (Xu hướng nhận diện) ──
        ws2 = wb.create_sheet(T["sheet_trends"])
        style_header(ws2["A1"], T["date"])
        style_header(ws2["B1"], T["scans_count"])
        ws2.row_dimensions[1].height = 20
        for idx, d in enumerate(daily_counts, 2):
            style_cell(ws2.cell(row=idx, column=1), d.get("date", ""), align="center")
            style_cell(ws2.cell(row=idx, column=2), d.get("count", 0), align="center")
            ws2.row_dimensions[idx].height = 20
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 24

        # ── Sheet 3: Top 5 Popular Breeds (Top 5 giống phổ biến) ──
        ws3 = wb.create_sheet(T["sheet_top5"])
        for col_idx, header_text in enumerate([T["rank"], T["breed"], T["scans_count"], T["pct"]], 1):
            style_header(ws3.cell(row=1, column=col_idx), header_text)
        ws3.row_dimensions[1].height = 20
        for idx, b in enumerate(top_breeds[:5], 2):
            pct_val = round(b.get("count", 0) / total_preds * 100, 1) if total_preds else 0.0
            breed_name = b.get("breed", "")
            if lang == "en":
                from i18n_server import translate_breed_vi_to_en
                breed_name = translate_breed_vi_to_en(breed_name)
            style_cell(ws3.cell(row=idx, column=1), idx - 1, align="center")
            style_cell(ws3.cell(row=idx, column=2), breed_name)
            style_cell(ws3.cell(row=idx, column=3), b.get("count", 0), align="center")
            style_cell(ws3.cell(row=idx, column=4), pct_val, align="center")
            ws3.row_dimensions[idx].height = 20
        ws3.column_dimensions["A"].width = 10
        ws3.column_dimensions["B"].width = 32
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 16

        # ── Sheet 4: Breed Distribution (Phân bố giống chó) ──
        ws4 = wb.create_sheet(T["sheet_dist"])
        for col_idx, header_text in enumerate([T["breed"], T["scans_count"], T["pct"]], 1):
            style_header(ws4.cell(row=1, column=col_idx), header_text)
        ws4.row_dimensions[1].height = 20
        for idx, b in enumerate(breed_dist, 2):
            style_cell(ws4.cell(row=idx, column=1), b["breed"])
            style_cell(ws4.cell(row=idx, column=2), b["count"], align="center")
            style_cell(ws4.cell(row=idx, column=3), b["pct"], align="center")
            ws4.row_dimensions[idx].height = 20
        ws4.column_dimensions["A"].width = 32
        ws4.column_dimensions["B"].width = 20
        ws4.column_dimensions["C"].width = 16

        # ── Sheet 5: Confidence Distribution (Phân bố độ tin cậy) ──
        ws5 = wb.create_sheet(T["sheet_conf"])
        style_header(ws5["A1"], T["conf_interval"])
        style_header(ws5["B1"], T["scans_count"])
        ws5.row_dimensions[1].height = 20
        for idx, (label, val) in enumerate(zip(conf_labels, confidence_dist), 2):
            style_cell(ws5.cell(row=idx, column=1), label, align="center")
            style_cell(ws5.cell(row=idx, column=2), val, align="center")
            ws5.row_dimensions[idx].height = 20
        ws5.column_dimensions["A"].width = 24
        ws5.column_dimensions["B"].width = 16

        # ── Sheet 6: Recent Results (Kết quả gần đây) ──
        ws6 = wb.create_sheet(T["sheet_recent"])
        for col_idx, header_text in enumerate([T["time"], T["breed"], T["species"], T["conf"]], 1):
            style_header(ws6.cell(row=1, column=col_idx), header_text)
        ws6.row_dimensions[1].height = 20
        for idx, p in enumerate(recent_predictions, 2):
            p_date = p["created_at"].strftime('%d/%m/%Y %H:%M') if p.get("created_at") else ""
            p_breed = p["breed"]
            if lang == "en":
                from i18n_server import translate_breed_vi_to_en
                p_breed = p.get("breed_en") or translate_breed_vi_to_en(p_breed)
            p_species = T["species_dog"] if p.get("species") in ("Dog", "Chó") else p.get("species", "")
            p_conf_val = f"{round((p.get('confidence') or 0) * 100)}%"
            
            style_cell(ws6.cell(row=idx, column=1), p_date, align="center")
            style_cell(ws6.cell(row=idx, column=2), p_breed)
            style_cell(ws6.cell(row=idx, column=3), p_species, align="center")
            style_cell(ws6.cell(row=idx, column=4), p_conf_val, align="center")
            ws6.row_dimensions[idx].height = 20
        ws6.column_dimensions["A"].width = 20
        ws6.column_dimensions["B"].width = 32
        ws6.column_dimensions["C"].width = 16
        ws6.column_dimensions["D"].width = 16

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{filename_base}.xlsx"
        )

    except Exception as e:
        print(f"Error exporting statistics: {e}")
        return jsonify({"error": "Xuất báo cáo thất bại"}), 500
    finally:
        if conn:
            conn.close()

